from flask import Blueprint, render_template, session, redirect, url_for, send_file, request, flash
from utils.exporter import export_players_excel, export_players_csv, export_players_html
from utils.db import get_db

export = Blueprint('export', __name__)


@export.route('/export')
def index():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    return render_template('export.html')


@export.route('/export/players/<fmt>')
def players(fmt):
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))

    if fmt == 'xlsx':
        data = export_players_excel()
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        extension = 'xlsx'
    elif fmt == 'csv':
        data = export_players_csv()
        mimetype = 'text/csv'
        extension = 'csv'
    elif fmt == 'html':
        data = export_players_html()
        mimetype = 'text/html'
        extension = 'html'
    else:
        return redirect(url_for('export.index'))

    return send_file(data, as_attachment=True,
                     download_name=f'aurus_players.{extension}',
                     mimetype=mimetype)


@export.route('/export/import', methods=['GET', 'POST'])
def import_excel():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    if session.get('access_level', 0) < 7:
        flash('Нет прав', 'danger')
        return redirect(url_for('export.index'))
    if request.method == 'POST':
        file = request.files.get('file')
        if file and file.filename.endswith(('.xlsx', '.xls')):
            filepath = f'data/{file.filename}'
            file.save(filepath)
            imported, skipped = import_players_from_excel(filepath)
            flash(f'Импортировано: {imported}, пропущено (дубли): {skipped}', 'success')
            return redirect(url_for('players.list'))
        flash('Неверный формат файла', 'danger')
    return render_template('import.html')


def import_players_from_excel(filepath):
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if 'Игроки' not in wb.sheetnames:
        return 0, 0
    ws = wb['Игроки']

    db = get_db()
    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1] or row[1] in ('Основной ник', None, ''):
            continue

        nick = str(row[1]).strip()
        if not nick:
            continue

        existing = db.execute('SELECT id FROM players WHERE nick = ?', (nick,)).fetchone()
        if existing:
            skipped += 1
            continue

        points = 0
        if row[9]:
            try:
                points = int(str(row[9]).replace(' ', '').replace(',', ''))
            except (ValueError, TypeError):
                points = 0

        db.execute(
            '''INSERT INTO players (nick, rank_in_game, points, rating1, rating2, rating3,
               planets, coordinates, registration_date, last_online, activity, player_status)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (nick, str(row[8]).strip() if row[8] else '',
             points, str(row[10]) if row[10] else '', str(row[11]) if row[11] else '',
             str(row[12]) if row[12] else '', str(row[13]) if row[13] else '',
             str(row[14]) if row[14] else '', str(row[15]) if row[15] else '',
             str(row[16]) if row[16] else '', str(row[17]) if row[17] else '',
             str(row[18]) if row[18] else '')
        )
        imported += 1

    db.commit()
    db.close()
    return imported, skipped
